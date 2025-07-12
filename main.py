import cv2
import pickle
import cvzone
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load gambar
image_path = 'p1.jpeg'
img_original = cv2.imread(image_path)
if img_original is None:
    raise FileNotFoundError(f"Gambar '{image_path}' tidak ditemukan.")

img = cv2.resize(img_original, (1080, 720))

# Load posisi parkir
try:
    with open('ParkirPos', 'rb') as f:
        posList = pickle.load(f)
except FileNotFoundError:
    raise FileNotFoundError("File 'ParkirPos' tidak ditemukan. Jalankan PemilihParkir.py terlebih dahulu untuk menentukan posisi parkir.")
except EOFError:
    posList = []  # Jika file kosong, gunakan daftar kosong

width, height = 50, 50

# Fungsi untuk menulis ke Excel
def update_excel(slot_data):
    file_name = "StatusParkirUPI.xlsx"
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Kode Slot", "Status", "Terakhir Diperbarui"])  # Header

    for i, (label, status) in enumerate(slot_data):
        if i + 2 > len(sheet['A']):  # Tambahkan baris baru jika belum ada
            sheet.append([label, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        else:
            sheet.cell(row=i + 2, column=1).value = label
            sheet.cell(row=i + 2, column=2).value = status
            sheet.cell(row=i + 2, column=3).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    wb.save(file_name)

# Fungsi untuk membuat label slot (A1, B1, ...)
def generate_label(index):
    row = index // 10
    col = index % 10
    return f"{chr(65 + row)}{col + 1}"

def checkSlot(imgPro):
    spaceCounter = 0
    slot_data = []

    for i, corners in enumerate(posList):
        if isinstance(corners, list) and all(isinstance(c, tuple) for c in corners):
            points = np.array(corners, np.int32)
        else:
            print(f"Error: Format data di 'ParkirPos' salah pada indeks {i}. Data: {corners}")
            continue

        label = generate_label(i)

        # Pastikan area kotak tidak keluar dari batas gambar
        x_min = min(pt[0] for pt in corners)
        y_min = min(pt[1] for pt in corners)
        x_max = max(pt[0] for pt in corners)
        y_max = max(pt[1] for pt in corners)

        if y_max > imgPro.shape[0] or x_max > imgPro.shape[1]:
            print(f"Warning: Kotak parkir {label} keluar dari batas gambar.")
            continue

        # Mask untuk mendeteksi isi kotak
        mask = np.zeros(imgPro.shape, dtype=np.uint8)
        cv2.fillPoly(mask, [points], 255)
        masked_img = cv2.bitwise_and(imgPro, mask)
        count = cv2.countNonZero(masked_img)

        # Status slot parkir
        if count < 300:
            color = (0, 255, 0)
            thickness = 3
            spaceCounter += 1
            status = "Kosong"
        else:
            color = (0, 0, 255)
            thickness = 2
            status = "Terisi"

        # Gambar kotak parkir
        cv2.polylines(img, [points], isClosed=True, color=color, thickness=thickness)
        cv2.putText(img, label, (x_min + 15, y_min + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255 ,255) , 2)
        slot_data.append((label, status))

    # Tampilkan jumlah slot parkir yang tersedia
    cvzone.putTextRect(img, f'Free: {spaceCounter}/{len(posList)}', (50, 50),
                       scale=2, thickness=3, offset=10, colorR=(0, 200, 0))
    
    # Perbarui data ke Excel
    update_excel(slot_data)


# Preprocessing gambar
imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
imgBlur = cv2.GaussianBlur(imgGray, (3, 3), 1)
imgThreshold = cv2.adaptiveThreshold(imgBlur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                     cv2.THRESH_BINARY_INV, 25, 16)
imgMedian = cv2.medianBlur(imgThreshold, 5)
kernel = np.ones((3, 3), np.int8)
imgDilate = cv2.dilate(imgMedian, kernel, iterations=1)

# Periksa slot parkir
checkSlot(imgDilate)

# Tampilkan hasil
cv2.namedWindow("Image", cv2.WINDOW_NORMAL)
cv2.imshow("Image", img)
#cv2.imshow("ImageBlur", imgBlur)
#cv2.imshow("ImageThres", imgThreshold)
cv2.waitKey(0)
cv2.destroyAllWindows()
